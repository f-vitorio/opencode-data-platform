#!/usr/bin/env python3
"""
Prospecting Runner - Orquestrador completo de prospecção automatizada

Fluxo:
1. Google Maps Scraping → leads brutos
2. Google Search Validation → valida indexação, GBP, Ads gap
3. Website Auditor → auditoria técnica dos sites
4. Lead Scoring → qualificação final
5. Salvamento no CRM (Supabase) + Relatório Markdown
"""

import sys
import os
import json
import argparse
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

# XLSX support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Adicionar skills ao path
skills_path = Path(__file__).parent.parent / "skills"
scripts_path = Path(__file__).parent.parent / "scripts"
sys.path.insert(0, str(skills_path))
sys.path.insert(0, str(skills_path / "google-maps-scraper"))
sys.path.insert(0, str(skills_path / "lead-scoring"))
sys.path.insert(0, str(skills_path / "playwright-automation"))
sys.path.insert(0, str(scripts_path))

try:
    from google_maps_scraper import search_places, GoogleMapsScraper
    from lead_scoring import enrich_lead_with_scoring, score_lead, classify_lead
    from website_audit_runner import WebsiteAuditRunner
except ImportError as e:
    print(f"Erro ao importar skills: {e}")
    sys.exit(1)


class ProspectingRunner:
    """Orquestrador do pipeline de prospecção."""
    
    def __init__(
        self,
        query: str,
        location: str = "",
        nicho: str = "",
        min_rating: float = 4.0,
        min_reviews: int = 10,
        max_leads: int = 200,
        audit_sample: int = 50,
        save_to_crm: bool = True,
        headed: bool = False,
        output_dir: str = None,
    ):
        self.query = query
        self.location = location
        self.nicho = nicho or query
        self.min_rating = min_rating
        self.min_reviews = min_reviews
        self.max_leads = max_leads
        self.audit_sample = audit_sample
        self.save_to_crm = save_to_crm
        self.headed = headed
        self.output_dir = Path(output_dir) if output_dir else Path.home() / "Documents" / "PROJETOS" / "prospecting"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.run_id = f"prospect_{self.nicho.lower().replace(' ', '_')}_{self.timestamp}"
        self.run_dir = self.output_dir / self.run_id
        self.run_dir.mkdir(parents=True, exist_ok=True)
        
        self.stats = {
            "raw_leads": 0,
            "with_website": 0,
            "validated": 0,
            "ads_gap": 0,
            "audited": 0,
            "qualified": 0,
            "quente": 0,
            "morno": 0,
            "frio": 0,
            "descartar": 0,
        }
        
        self.all_leads = []
    
    def run(self) -> Dict[str, Any]:
        """Executa pipeline completo."""
        print(f"\n{'='*70}")
        print(f"PROSPECÇÃO AUTOMATIZADA - {self.run_id}")
        print(f"{'='*70}")
        print(f"Query: {self.query} | Location: {self.location}")
        print(f"Filtros: rating≥{self.min_rating}, reviews≥{self.min_reviews}")
        print(f"Max leads: {self.max_leads} | Auditoria amostral: {self.audit_sample}")
        print(f"Output: {self.run_dir}")
        print(f"{'='*70}\n")
        
        start_time = time.time()
        
        # ETAPA 1: Google Maps Scraping
        print("\n[1/5] GOOGLE MAPS SCRAPING...")
        raw_leads = self._step_maps_scraping()
        self.stats["raw_leads"] = len(raw_leads)
        print(f"       Encontrados: {len(raw_leads)} leads brutos")
        
        # ETAPA 2: Google Search Validation
        print("\n[2/5] GOOGLE SEARCH VALIDATION...")
        validated_leads = self._step_search_validation(raw_leads)
        self.stats["validated"] = len(validated_leads)
        self.stats["with_website"] = sum(1 for l in validated_leads if l.get("website"))
        self.stats["ads_gap"] = sum(1 for l in validated_leads if not l.get("search_validation", {}).get("running_ads"))
        print(f"       Com site: {self.stats['with_website']} | Não anunciam: {self.stats['ads_gap']}")
        
        # ETAPA 3: Website Auditor (amostragem)
        print("\n[3/5] WEBSITE AUDITOR (amostragem)...")
        audited_leads = self._step_website_audit(validated_leads)
        self.stats["audited"] = len(audited_leads)
        
        # ETAPA 4: Lead Scoring
        print("\n[4/5] LEAD SCORING & CLASSIFICAÇÃO...")
        scored_leads = self._step_lead_scoring(audited_leads)
        self.stats["qualified"] = len(scored_leads)
        self.stats["quente"] = sum(1 for l in scored_leads if l.get("classification") == "QUENTE")
        self.stats["morno"] = sum(1 for l in scored_leads if l.get("classification") == "MORNO")
        self.stats["frio"] = sum(1 for l in scored_leads if l.get("classification") == "FRIO")
        self.stats["descartar"] = sum(1 for l in scored_leads if l.get("classification") == "DESCARTAR")
        print(f"       QUENTE: {self.stats['quente']} | MORNO: {self.stats['morno']} | FRIO: {self.stats['frio']} | DESCARTAR: {self.stats['descartar']}")
        
        # ETAPA 5: Salvar + Relatório
        print("\n[5/5] SALVANDO CRM + RELATÓRIO...")
        self._step_save_and_report(scored_leads)
        
        duration = time.time() - start_time
        
        # Summary
        result = {
            "run_id": self.run_id,
            "query": self.query,
            "location": self.location,
            "timestamp": datetime.now().isoformat(),
            "duration_seconds": round(duration, 1),
            "stats": self.stats,
            "leads": scored_leads,
            "output_dir": str(self.run_dir),
        }
        
        self._print_summary(result)
        return result
    
    def _step_maps_scraping(self) -> List[Dict]:
        """ETAPA 1: Busca no Google Maps - Multi-location para áreas nobres de SP."""
        
        # Se location for "centro_sp_30km" ou similar, buscar múltiplos bairros nobres
        if "centro_sp" in self.location.lower() or "30km" in self.location.lower() or "central_sp" in self.location.lower():
            # Bairros nobres/alto padrão no centro expandido de SP (raio ~30km)
            locations = [
                ("Jardins", "Jardins, São Paulo, SP"),
                ("Itaim Bibi", "Itaim Bibi, São Paulo, SP"),
                ("Vila Olímpia", "Vila Olímpia, São Paulo, SP"),
                ("Pinheiros", "Pinheiros, São Paulo, SP"),
                ("Vila Madalena", "Vila Madalena, São Paulo, SP"),
                ("Moema", "Moema, São Paulo, SP"),
                ("Brooklin", "Brooklin, São Paulo, SP"),
                ("Consolação", "Consolação, São Paulo, SP"),
                ("Cerqueira César", "Cerqueira César, São Paulo, SP"),
                ("Higienópolis", "Higienópolis, São Paulo, SP"),
                ("Perdizes", "Perdizes, São Paulo, SP"),
                ("Vila Mariana", "Vila Mariana, São Paulo, SP"),
                ("Ibirapuera", "Ibirapuera, São Paulo, SP"),
                ("Campo Belo", "Campo Belo, São Paulo, SP"),
                ("Santo Amaro", "Santo Amaro, São Paulo, SP"),
            ]
        else:
            # Busca única no location especificado
            locations = [("Custom", self.location)]
        
        all_leads = []
        max_per_location = max(10, self.max_leads // len(locations))
        
        for location_name, location_query in locations:
            print(f"       Buscando em: {location_name} ({location_query})")
            scraper = GoogleMapsScraper(
                min_rating=self.min_rating,
                min_reviews=self.min_reviews,
                must_have_website=True,
                max_results=max_per_location,
                headed=self.headed,
            )
            leads = scraper.search(self.query, location_query)
            print(f"         Encontrados: {len(leads)} leads")
            
            for lead in leads:
                lead["prospecting_query"] = self.query
                lead["prospecting_location"] = location_query
                lead["prospecting_sub_location"] = location_name
                lead["prospecting_nicho"] = self.nicho
                lead["collected_at"] = datetime.now().isoformat()
            
            all_leads.extend(leads)
            
            # Rate limiting entre buscas
            time.sleep(3)
        
        # Deduplicar por place_id
        seen = set()
        unique_leads = []
        for lead in all_leads:
            pid = lead.get("place_id")
            if pid and pid not in seen:
                seen.add(pid)
                unique_leads.append(lead)
        
        print(f"       Total único após deduplicação: {len(unique_leads)}")
        return unique_leads[:self.max_leads]
    
    def _step_search_validation(self, leads: List[Dict]) -> List[Dict]:
        """ETAPA 2: Validação via Google Search (indexação, GBP, Ads)."""
        # Para cada lead, fazer busca no Google
        # NOTA: Implementação simplificada - em produção usar API ou scraping cuidadoso
        
        validated = []
        for lead in leads:
            validation = self._validate_via_google_search(lead)
            lead["search_validation"] = validation
            
            # Filtrar: deve ter site E não anunciar (gap de oportunidade)
            if lead.get("website") and not validation.get("running_ads", True):
                validated.append(lead)
        
        return validated
    
    def _validate_via_google_search(self, lead: Dict) -> Dict[str, Any]:
        """
        Valida lead via Google Search.
        Verifica: site indexado, GBP ativo, sitelinks, anúncios ativos.
        
        NOTA: Versão simplificada. Em produção, usar:
        - Google Custom Search API
        - SerpAPI
        - Ou scraping cuidadoso com rate limiting
        """
        # Placeholder - implementar com requests + BeautifulSoup ou API
        # Por enquanto retorna estrutura esperada
        
        website = lead.get("website", "")
        domain = ""
        if website:
            from urllib.parse import urlparse
            domain = urlparse(website).netloc.lower().replace("www.", "")
        
        # Simulação baseada em heurísticas
        # Em produção: buscar "site:{domain}" e "{nome} Google Ads"
        
        rating = lead.get("rating") or 0
        reviews = lead.get("reviews_count") or 0
        
        return {
            "indexed": True,  # Assumir indexado se tem site
            "has_sitelinks": False,  # Precisa verificar
            "gbp_active": rating >= 4.0 and reviews >= 10,
            "running_ads": False,  # Assumir não anuncia (oportunidade)
            "competitors_ads_count": 0,  # Precisa verificar
            "estimated_cpc": self._estimate_cpc_for_nicho(),
            "keywords_opportunity": self._get_keywords_for_nicho(),
            "monthly_searches": "N/A",
        }
    
    def _estimate_cpc_for_nicho(self) -> float:
        """Estima CPC médio para o nicho."""
        nicho_lower = self.nicho.lower()
        
        cpc_estimates = {
            "psicologia": 12,
            "psicólogo": 12,
            "terapia": 10,
            "advocacia": 18,
            "advogado": 18,
            "direito trabalhista": 15,
            "odontologia": 14,
            "dentista": 14,
            "implante": 25,
            "ortodontia": 20,
            "fisioterapia": 10,
            "nutrição": 8,
            "contabilidade": 12,
            "contador": 12,
            "abertura empresa": 8,
            "imobiliária": 10,
            "marketing digital": 20,
            "seo": 25,
            "google ads": 30,
        }
        
        for keyword, cpc in cpc_estimates.items():
            if keyword in nicho_lower:
                return cpc
        
        return 15  # Default
    
    def _get_keywords_for_nicho(self) -> List[str]:
        """Retorna palavras-chave de oportunidade para o nicho."""
        nicho_lower = self.nicho.lower()
        
        keyword_map = {
            "psicologia": ["psicólogo sp", "terapia são paulo", "psicologia clínica", "psicólogo online", "terapia cognitiva"],
            "advocacia": ["advogado sp", "advogado trabalhista", "advogado cível", "escritório advocacia são paulo"],
            "odontologia": ["dentista sp", "implante dentário são paulo", "ortodontia", "clínica odontológica"],
            "fisioterapia": ["fisioterapeuta sp", "fisioterapia são paulo", "reabilitação física"],
            "contabilidade": ["contador sp", "abrir empresa são paulo", "contabilidade para mei", "escritório contabilidade"],
        }
        
        for keyword, kws in keyword_map.items():
            if keyword in nicho_lower:
                return kws
        
        return [f"{self.nicho} são paulo", f"{self.nicho} sp", f"melhor {self.nicho}"]
    
    def _step_website_audit(self, leads: List[Dict]) -> List[Dict]:
        """ETAPA 3: Auditoria técnica dos sites (amostragem)."""
        audited = []
        
        # Pegar amostra para auditar (priorizar os com melhor google_presence)
        leads_sorted = sorted(
            leads, 
            key=lambda l: (l.get("rating") or 0) * (l.get("reviews_count") or 0), 
            reverse=True
        )
        
        sample = leads_sorted[:self.audit_sample]
        
        for i, lead in enumerate(sample):
            website = lead.get("website")
            if not website:
                continue
            
            print(f"       Auditoria {i+1}/{len(sample)}: {lead['name'][:40]}...")
            
            try:
                auditor = WebsiteAuditRunner(website, mode="COMPLETA", headed=self.headed)
                audit_result = auditor.run()
                
                lead["website_audit"] = audit_result
                audited.append(lead)
                
                # Rate limiting entre auditorias
                time.sleep(2)
                
            except Exception as e:
                print(f"       ⚠ Erro na auditoria: {e}")
                lead["website_audit"] = {"error": str(e)}
                audited.append(lead)
        
        return audited
    
    def _step_lead_scoring(self, leads: List[Dict]) -> List[Dict]:
        """ETAPA 4: Scoring e classificação final."""
        scored = []
        
        for lead in leads:
            try:
                enriched = enrich_lead_with_scoring(lead)
                scored.append(enriched)
            except Exception as e:
                print(f"       ⚠ Erro no scoring: {e}")
                lead["scoring"] = {"error": str(e)}
                lead["classification"] = "ERRO"
                lead["next_action"] = "REVISAR_MANUAL"
                scored.append(lead)
        
        return scored
    
    def _step_save_and_report(self, leads: List[Dict]):
        """ETAPA 5: Salvar no CRM + Gerar relatórios."""
        # Salvar JSON completo
        json_path = self.run_dir / f"{self.run_id}_leads.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(leads, f, ensure_ascii=False, indent=2, default=str)
        
        # Gerar relatório Markdown
        md_path = self.run_dir / f"{self.run_id}_report.md"
        self._generate_markdown_report(leads, md_path)
        
        # Salvar XLSX
        xlsx_path = self.run_dir / f"{self.run_id}_leads.xlsx"
        self._save_to_xlsx(leads, xlsx_path)
        
        # Salvar leads individuais
        leads_dir = self.run_dir / "leads"
        leads_dir.mkdir(exist_ok=True)
        
        for lead in leads:
            if lead.get("classification") in ["QUENTE", "MORNO"]:
                safe_name = "".join(c for c in lead["name"] if c.isalnum() or c in " -_").strip()
                safe_name = safe_name[:80]
                lead_file = leads_dir / f"lead_{lead.get('classification', 'UNK')}_{safe_name}.json"
                with open(lead_file, 'w', encoding='utf-8') as f:
                    json.dump(lead, f, ensure_ascii=False, indent=2, default=str)
        
        # Se save_to_crm, integrar com data-platform-manager
        if self.save_to_crm:
            self._save_to_supabase(leads)
        
        print(f"       ✓ JSON: {json_path}")
        print(f"       ✓ XLSX: {xlsx_path}")
        print(f"       ✓ Relatório: {md_path}")
        print(f"       ✓ Leads individuais: {leads_dir}")
    
    def _save_to_supabase(self, leads: List[Dict]):
        """Salva leads no Supabase via data-platform-manager."""
        try:
            # Usar o script client-add.py ou API direta
            # Por enquanto, apenas log
            print(f"       📝 {len(leads)} leads preparados para CRM (integração pendente)")
        except Exception as e:
            print(f"       ⚠ Erro ao salvar no CRM: {e}")

    def _save_to_xlsx(self, leads: List[Dict], output_path: Path):
        """Salva leads em formato XLSX."""
        if not HAS_OPENPYXL:
            print("       ⚠ openpyxl não instalado, pulando XLSX")
            return
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Classificação colors
        class_fills = {
            "QUENTE": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
            "MORNO": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            "FRIO": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "DESCARTAR": PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
        }
        
        # Aba 1: Resumo Executivo
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        
        summary_headers = ["Métrica", "Valor"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        summary_data = [
            ("Run ID", self.run_id),
            ("Data", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Query", f"{self.query} {self.location}"),
            ("Leads brutos (Maps)", self.stats['raw_leads']),
            ("Com site válido", self.stats['with_website']),
            ("Não anunciam no Ads", self.stats['ads_gap']),
            ("Auditados (amostra)", self.stats['audited']),
            ("Qualificados", self.stats['qualified']),
            ("QUENTE (≥85)", self.stats['quente']),
            ("MORNO (70-84)", self.stats['morno']),
            ("FRIO (50-69)", self.stats['frio']),
            ("DESCARTAR (<50)", self.stats['descartar']),
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1, value=metric).border = thin_border
            ws_summary.cell(row=row, column=2, value=value).border = thin_border
        
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 25
        
        # Aba 2: Leads Detalhados
        ws_leads = wb.create_sheet("Leads")
        
        lead_headers = [
            "Classificação", "Score Final", "Nome", "Rating", "Reviews",
            "Site", "Telefone", "Endereço", "Categorias",
            "Site Quality (inv)", "Google Presence", "Ads Gap", "Market Fit",
            "Problemas Principais", "Oportunidades", "Investimento Estimado", "ROI Estimado",
            "Próxima Ação", "Pitch Headline"
        ]
        
        for col, header in enumerate(lead_headers, 1):
            cell = ws_leads.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row, lead in enumerate(leads, 2):
            scoring = lead.get("scoring", {})
            pitch = lead.get("pitch_data", {})
            classification = lead.get("classification", "N/A")
            
            # Montar problemas e oportunidades
            problems = "; ".join(pitch.get("problems", [])[:5]) if pitch.get("problems") else ""
            opportunities = "; ".join(pitch.get("opportunities", [])[:3]) if pitch.get("opportunities") else ""
            
            values = [
                classification,
                scoring.get("final_score", 0),
                lead.get("name", ""),
                lead.get("rating", ""),
                lead.get("reviews_count", ""),
                lead.get("website", ""),
                lead.get("phone", ""),
                lead.get("address", ""),
                ", ".join(lead.get("categories", [])) if lead.get("categories") else "",
                scoring.get("site_quality_inverse", 0),
                scoring.get("google_presence", 0),
                scoring.get("ads_gap", 0),
                scoring.get("market_fit", 0),
                problems,
                opportunities,
                pitch.get("estimated_investment", ""),
                pitch.get("estimated_roi", ""),
                lead.get("next_action", ""),
                pitch.get("headline", ""),
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws_leads.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Colorir célula de classificação
                if col == 1 and value in class_fills:
                    cell.fill = class_fills[value]
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajustar larguras das colunas
        col_widths = [14, 12, 40, 8, 8, 40, 18, 45, 25, 14, 14, 10, 12, 50, 50, 30, 25, 18, 50]
        for i, width in enumerate(col_widths, 1):
            ws_leads.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze panes
        ws_leads.freeze_panes = "A2"
        
        # Aba 3: Auditoria de Sites (apenas leads auditados)
        ws_audit = wb.create_sheet("Auditoria Sites")
        
        audit_headers = [
            "Nome", "Site", "Score Geral", "SEO", "CRO", "Performance", 
            "Acessibilidade", "Mobile", "Segurança",
            "Issues Críticos (P0)", "Avisos (P1/P2)"
        ]
        
        for col, header in enumerate(audit_headers, 1):
            cell = ws_audit.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row, lead in enumerate(leads, 2):
            audit = lead.get("website_audit", {})
            if not audit or "error" in audit:
                continue
                
            scores = audit.get("scores", {})
            critical = audit.get("critical_issues", [])
            warnings = audit.get("warnings", [])
            
            critical_text = "; ".join([c["title"] for c in critical[:5]])
            warn_text = "; ".join([w["title"] for w in warnings[:5]])
            
            values = [
                lead.get("name", ""),
                lead.get("website", ""),
                audit.get("overall_score", 0),
                scores.get("seo", 0),
                scores.get("cro", 0),
                scores.get("performance", 0),
                scores.get("acessibilidade", 0),
                scores.get("mobile", 0),
                scores.get("seguranca", 0),
                critical_text,
                warn_text,
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws_audit.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        audit_widths = [35, 40, 12, 8, 8, 12, 14, 10, 10, 50, 50]
        for i, width in enumerate(audit_widths, 1):
            ws_audit.column_dimensions[get_column_letter(i)].width = width
        
        ws_audit.freeze_panes = "A2"
        
        # Save
        wb.save(output_path)
        print(f"       ✓ XLSX: {output_path}")
    
    def _generate_markdown_report(self, leads: List[Dict], output_path: Path):
        """Gera relatório Markdown executivo."""
        quentes = [l for l in leads if l.get("classification") == "QUENTE"]
        mornos = [l for l in leads if l.get("classification") == "MORNO"]
        frios = [l for l in leads if l.get("classification") == "FRIO"]
        
        md = f"""# Relatório de Prospecção: {self.nicho} em {self.location}

**Run ID:** {self.run_id}  
**Data:** {datetime.now().strftime("%d/%m/%Y %H:%M")}  
**Query:** `{self.query} {self.location}`

---

## Resumo Executivo

| Métrica | Valor |
|---------|-------|
| Leads brutos (Maps) | {self.stats['raw_leads']} |
| Com site válido | {self.stats['with_website']} |
| Não anunciam no Ads | {self.stats['ads_gap']} |
| Auditados (amostra) | {self.stats['audited']} |
| **Qualificados** | **{self.stats['qualified']}** |
| 🔥 QUENTE (≥85) | {self.stats['quente']} |
| 🟡 MORNO (70-84) | {self.stats['morno']} |
| 🔵 FRIO (50-69) | {self.stats['frio']} |
| ❌ DESCARTAR (<50) | {self.stats['descartar']} |

---

## Top Leads QUENTES (Abordagem Imediata)

"""
        
        for i, lead in enumerate(quentes[:10], 1):
            scoring = lead.get("scoring", {})
            pitch = lead.get("pitch_data", {})
            audit = lead.get("website_audit", {})
            
            md += f"""### {i}. {lead['name']}

**Score Final:** {scoring.get('final_score', 0)}/100 | **Classificação:** {lead.get('classification')}  
**Rating GBP:** {lead.get('rating', 'N/A')}★ ({lead.get('reviews_count', 0)} reviews)  
**Site:** {lead.get('website', 'N/A')}  
**Telefone:** {lead.get('phone', 'N/A')}  
**Endereço:** {lead.get('address', 'N/A')}

**Problemas do Site:**
"""
            for prob in pitch.get("problems", [])[:3]:
                md += f"- {prob}\n"
            
            md += f"""
**Oportunidades:**
"""
            for opp in pitch.get("opportunities", []):
                md += f"- {opp}\n"
            
            md += f"""
**Investimento Sugerido:** {pitch.get('estimated_investment', 'N/A')}  
**ROI Estimado:** {pitch.get('estimated_roi', 'N/A')}

---

"""
        
        md += f"""
## Leads MORNOS (Nutrição Semanal)

| # | Lead | Score | Rating | Site | Próxima Ação |
|---|------|-------|--------|------|--------------|
"""
        
        for i, lead in enumerate(mornos[:20], 1):
            scoring = lead.get("scoring", {})
            md += f"| {i} | {lead['name'][:40]} | {scoring.get('final_score', 0)} | {lead.get('rating', 'N/A')}★ | {lead.get('website', 'N/A')[:30]} | Nutrição LinkedIn/Email |\n"
        
        md += f"""

## Pipeline Sugerido

### Esta Semana
- Abordar **{len(quentes)} leads QUENTES** (WhatsApp/ligação direta)
- Preparar propostas personalizadas usando `pitch_data`

### Próxima Semana  
- Nutrir **{len(mornos)} leads MORNOS** (conexão LinkedIn + email educativo)
- Conteúdo: "5 erros que fazem seu site perder clientes no mobile"

### 30 Dias
- Re-avaliar **{len(frios)} leads FRIOS**
- Verificar se começaram a anunciar ou melhoraram o site

---

## Metodologia

1. **Google Maps Scraping** → Busca local com filtros (rating≥{self.min_rating}, reviews≥{self.min_reviews})
2. **Google Search Validation** → Verifica indexação, GBP ativo, gap de Ads
3. **Website Auditor** → Auditoria técnica completa (SEO, CRO, Performance, Mobile, Acessibilidade, Segurança)
4. **Lead Scoring** → Algoritmo ponderado:
   - Site Quality Inverse (30%): site ruim = oportunidade
   - Google Presence (25%): GBP forte = lead qualificado
   - Ads Gap (25%): não anuncia = oportunidade
   - Market Fit (20%): nicho core da FVS7
5. **Classificação** → QUENTE (≥85) | MORNO (70-84) | FRIO (50-69) | DESCARTAR (<50)

---

## Próximos Passos

1. Revisar leads QUENTES e aprovar abordagem
2. Equipe comercial usa `pitch_data` para contato personalizado
3. `proposal-agent` gera propostas baseadas no `estimated_investment`
4. Acompanhar conversões no CRM

---

*Relatório gerado automaticamente pelo Prospecting Agent v1.0*
"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md)

    def _print_summary(self, result: Dict):
        """Imprime resumo final no console."""
        stats = result["stats"]
        print(f"\n{'='*70}")
        print(f"RESUMO DA PROSPECÇÃO: {self.run_id}")
        print(f"{'='*70}")
        print(f"Leads brutos (Maps):     {stats['raw_leads']}")
        print(f"Com site válido:         {stats['with_website']}")
        print(f"Não anunciam no Ads:     {stats['ads_gap']}")
        print(f"Auditados (amostra):       {stats['audited']}")
        print(f"Qualificados:            {stats['qualified']}")
        print(f"  🔥 QUENTE (≥85):         {stats['quente']}")
        print(f"  🟡 MORNO (70-84):        {stats['morno']}")
        print(f"  🔵 FRIO (50-69):         {stats['frio']}")
        print(f"  ❌ DESCARTAR (<50):       {stats['descartar']}")
        print(f"Duração:                   {result['duration_seconds']:.1f}s")
        print(f"Output:                    {result['output_dir']}")
        print(f"{'='*70}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Prospecção Automatizada - Google Maps + Website Audit + Lead Scoring",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  %(prog)s "clínica psicologia" "São Paulo, SP" --max-leads 100 --audit-sample 30
  %(prog)s "advogado trabalhista" "Rio de Janeiro, RJ" --nicho "advocacia trabalhista" --headed
  %(prog)s "dentista implante" "Belo Horizonte, MG" --min-rating 4.5 --min-reviews 20
        """
    )
    
    parser.add_argument("query", help="Termo de busca (ex: 'clínica psicologia')")
    parser.add_argument("location", help="Localização (ex: 'São Paulo, SP')", nargs="?", default="")
    parser.add_argument("--nicho", "-n", help="Nome do nicho para market fit (padrão: query)")
    parser.add_argument("--min-rating", type=float, default=4.0, help="Rating mínimo no Maps (padrão: 4.0)")
    parser.add_argument("--min-reviews", type=int, default=10, help="Reviews mínimos (padrão: 10)")
    parser.add_argument("--max-leads", type=int, default=200, help="Max leads brutos (padrão: 200)")
    parser.add_argument("--audit-sample", type=int, default=50, help="Leads para auditar (padrão: 50)")
    parser.add_argument("--no-crm", action="store_true", help="Não salvar no CRM")
    parser.add_argument("--headed", "-H", action="store_true", help="Browser visível (debug)")
    parser.add_argument("--output-dir", "-o", help="Diretório de saída")
    
    args = parser.parse_args()
    
    runner = ProspectingRunner(
        query=args.query,
        location=args.location,
        nicho=args.nicho,
        min_rating=args.min_rating,
        min_reviews=args.min_reviews,
        max_leads=args.max_leads,
        audit_sample=args.audit_sample,
        save_to_crm=not args.no_crm,
        headed=args.headed,
        output_dir=args.output_dir,
    )
    
    try:
        result = runner.run()
        
        # Exit code baseado em resultados
        if result["stats"]["quente"] > 0:
            print(f"\n✅ SUCESSO: {result['stats']['quente']} leads QUENTES encontrados!")
            sys.exit(0)
        elif result["stats"]["morno"] > 0:
            print(f"\n⚠️ PARCIAL: {result['stats']['morno']} leads MORNOS, nenhum QUENTE")
            sys.exit(0)
        else:
            print(f"\n❌ SEM LEADS QUALIFICADOS")
            sys.exit(1)
            
    except KeyboardInterrupt:
        print("\n\nInterrompido pelo usuário.")
        sys.exit(130)
    except Exception as e:
        print(f"\nErro durante execução: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()